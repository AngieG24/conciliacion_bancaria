import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import re

# -----------------------------------------------------------------------
#                             Conciliación bancaria
#  ----------------------------------------------------------------------

# 1. Reglas de trasnformación segun el tipo de fuente 

reglas_fuente = {
        "Extracto": {
            "columnas": {
                "Cuenta": 0,
                "Fecha": 2,
                "Descripcion": 3,
                "Importe": 4,
                "Ref 1": 5,
                "Ref 2": 6,
            },

        },

            "Auxiliar": {
                "columnas": {
                    "Cuenta":8,
                    "Fecha": 4,
                    "Asiento": 5,
                    "Descripcion": 10,
                    "Importe": 15,
                    "Nid": 18,
                    "IT Tercero": 16,
                    "Tercero": 17,
                    "Responsable": 29,
                    "Fecha_hora modificacion": 31

        },
    }

}

# 2. Función para parsear múltiples formatos de fecha

formatos_fecha = [
    "%d.%m.%y",
    "%d.%m.%Y",
    "%d/%m/%Y",
    "%Y%m%d",
    "%Y-%m-%d",
    "%d-%m-%Y"
]

def parsear_fecha_multiple_conciliacion(valor):
    """Intenta convertir un valor a fecha probando varios formatos."""
    if pd.isna(valor):              # Verifica si el valor está vacío o es nulo
        return pd.NaT               # Si el valor es nulo (por ejemplo, está vacío o es NaN), devuelve pd.NaT (que significa "Not a Time", es decir, no hay fecha).
    valor = str(valor).strip()      # Convierte el valor a texto y elimina espacios en blanco al inicio y al final
    for formato in formatos_fecha:  # Intenta convertir el valor a fecha usando varios formatos
        try:
            return pd.to_datetime(valor, format=formato, errors="raise")
        except:
            continue
    return pd.NaT

# 3. Función genérica de transformación para conciliación bancaria
    
def transformar_extracto_conciliacion(df, fuente):
    
    if fuente not in reglas_fuente:
        raise ValueError(f"No hay reglas definidas para la fuente '{fuente}'")
    
    
    reglas = reglas_fuente.get(fuente)
    columnas = reglas['columnas']
  
    df_out_conciliacion = df.copy()

    
    # ✅ Columna: Tipo

    df_out_conciliacion['Tipo'] = fuente

    # ✅ Columna: Cuenta

    df_out_conciliacion['Cuenta'] = df.iloc[:, columnas['Cuenta']].astype(str).str.strip()

    # ✅ Columna: Fecha
    df_out_conciliacion['Fecha'] = (
        df.iloc[:, columnas['Fecha']]
        .astype(str)
        .str.strip()
        .apply(parsear_fecha_multiple_conciliacion)
    )    

    # ✅ Columna: Descripción

    df_out_conciliacion['Descripcion'] = df.iloc[:, columnas['Descripcion']].astype(str).str.strip()

    # ✅ Columna: importe

    df_out_conciliacion['Importe'] = pd.to_numeric(df.iloc[:, columnas['Importe']],errors="coerce").fillna(0)

    # ✅ Columna: Importe 2 (ajuste de signo para Auxiliar)
    if fuente == "Auxiliar":
        df_out_conciliacion['Importe 2'] = df_out_conciliacion['Importe'] * -1
    else:
        df_out_conciliacion['Importe 2'] = df_out_conciliacion['Importe']

    # ✅ Columnas opcionales

    opcionales_extracto = {
        'Asiento': lambda s: s.fillna("").astype(str).str.lstrip('0').str.upper(),
        'Nid': lambda s: s.fillna("").astype(str).str.lstrip('0').str.upper(),
        'IT Tercero': lambda s: s.fillna("").astype(str).str.upper(),
        'Tercero': lambda s: s.fillna("").astype(str).str.strip().str.upper(),
        'Responsable': lambda s: s.fillna("").astype(str).str.strip().str.upper(),
        'Fecha_hora modificacion': lambda s: s.fillna("").astype(str).str.strip(),
        'Ref 1': lambda s: s.fillna("").astype(str).str.upper(),
        'Ref 2': lambda s: s.fillna("").astype(str).str.lstrip('0').str.upper(),
    }

    for col, func in opcionales_extracto.items():
        if col in columnas:
            df_out_conciliacion[col] = func(df.iloc[:, columnas[col]])
        else:
            df_out_conciliacion[col] = ""


    df_final_conciliacion = df_out_conciliacion[['Tipo','Cuenta','Fecha', 'Asiento', 'Descripcion', 'Importe', 'Importe 2','Ref 1', 'Ref 2', 'Nid', 'IT Tercero', 'Tercero', 'Responsable', 'Fecha_hora modificacion']]
    return df_final_conciliacion

# 4. Reglas de cruce 

def conciliar_extractos(df_extracto, df_auxiliar):
    conciliados = []
    no_conciliados = []
    usados_aux = set()

    for i, row_ext in df_extracto.iterrows():
        encontrado = False

        for j, row_aux in df_auxiliar.iterrows():
            if j in usados_aux:
                continue

            # --- Reglas de cruce ---
            if (row_ext["Fecha"] == row_aux["Fecha"]) and \
               (row_ext["Descripcion"] == row_aux["Descripcion"]) and \
               (row_ext["Importe"] == row_aux["Importe"]):
                pass
            elif (row_ext["Fecha"] == row_aux["Fecha"]) and \
                 (row_ext["Importe"] == row_aux["Importe"]):
                pass
            elif abs((pd.to_datetime(row_ext["Fecha"], dayfirst=True) -
                      pd.to_datetime(row_aux["Fecha"], dayfirst=True)).days) <= 1 and \
                 (row_ext["Importe"] == row_aux["Importe"]):
                pass

            else:
                continue  # no match

            # ✅ Guardar match (conciliados)
            conciliados.append(row_ext.to_dict())
            conciliados.append(row_aux.to_dict())

            usados_aux.add(j)
            encontrado = True
            break

        if not encontrado:
            # Extracto sin match → guardarlo en bloque aparte
            no_conciliados.append(row_ext.to_dict())

    # Auxiliares no conciliados
    for j, row_aux in df_auxiliar.iterrows():
        if j not in usados_aux:
            no_conciliados.append(row_aux.to_dict())

    # Concatenar conciliados arriba + fila en blanco + no conciliados
    if conciliados and no_conciliados:
        conciliados.append({col: "" for col in df_extracto.columns})  # separador
    return pd.DataFrame(conciliados + no_conciliados)

# ----------------------------------- Interfaz Streamlit para conciliación bancaria -----------------------------------

# 🔹 Interfaz en Streamlit
st.title("Conciliación Bancaria")

col1, col2 = st.columns(2)

with col1:
    file_extracto = st.file_uploader("Sube archivo de Extracto", type=["xlsx"])
with col2:
    file_auxiliar = st.file_uploader("Sube archivo de Auxiliar", type=["xlsx"])

if file_extracto and file_auxiliar:
    df_ext_raw = pd.read_excel(file_extracto, header=None, skiprows=1)
    df_aux_raw = pd.read_excel(file_auxiliar, header=None, skiprows=1)

    df_ext = transformar_extracto_conciliacion(df_ext_raw, "Extracto")
    df_aux = transformar_extracto_conciliacion(df_aux_raw, "Auxiliar")

    st.subheader("Vista previa - Extracto")
    st.dataframe(df_ext.head(5))

    st.subheader("Vista previa - Auxiliar")
    st.dataframe(df_aux.head(5))

    # 🔹 Aplicar conciliación
    df_conciliado = conciliar_extractos(df_ext, df_aux)

    st.subheader("Resultado de la conciliación")
    st.dataframe(df_conciliado)

    # Descargar en Excel
    output_path = "conciliacion.xlsx"
    df_conciliado.to_excel(output_path, index=False, sheet_name="Conciliacion")
    

    # 🔹 Aplicar formato con openpyxl
    
    wb = load_workbook(output_path)
    ws = wb["Conciliacion"]

    # Buscar la columna "Importe" y "Tipo"
    col_importes = []
    for idx, cell in enumerate(ws[1], 1):  # Fila de encabezados
        if cell.value in ["Importe", "Importe 2"]:
            col_importes.append(idx)
        if cell.value == "Tipo":
            col_tipo = idx

    # Formato para columna "Importe"
    for col in col_importes:
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                # 🔹 Formato: miles con ".", decimales con ",", negativos en rojo
                cell.number_format = '#,##0.00;[Red]-#,##0.00'

    # Color de fondo para filas con Tipo = Extracto
    fill_extracto = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")

    if col_tipo:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[col_tipo - 1].value == "Extracto":
                for cell in row:
                    cell.fill = fill_extracto


    wb.save("conciliacion.xlsx")

    with open("conciliacion.xlsx", "rb") as f:
        st.download_button("📥 Descargar conciliación en Excel", f, file_name="conciliacion.xlsx")